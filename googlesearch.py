"""
/*******************************************************
 * General Public License 2024 
 
 * Authors:
 *   - Janaka Wijekoon
 *   - Rashini Liyanarachchi

 * Feel free to make any modifications to the code. If you make any changes, kindly add your name to the authors list.
 *******************************************************/

"""

import subprocess

def install_module(module_name):
    subprocess.check_call(["pip", "install", module_name])

# List the required modules
required_modules = ['webbrowser', 'openpyxl', 'urllib.parse']

# Try importing the modules and install the missing ones
for module in required_modules:
    try:
        __import__(module)
    except ImportError:
        print(f"{module} is not installed. Installing...")
        install_module(module)

import webbrowser
from openpyxl import load_workbook
from urllib.parse import quote

wb = load_workbook('dataset.xlsx')
sheet = wb.active # Get the active sheet. I had only one sheet in the file and copied all the titles to col A. 

#Get the starting records. 
start_record = int(input("Enter the number where to Begin?: "))  # Adjusting to zero-based index


# For each cell perform Google search
for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
    if i < start_record:
        continue
    for cell_value in row:
        if isinstance(cell_value, str):  # Check if cell contains text
            query = cell_value
            print("=============================")
            print(f"Search query: '{query}'")
            confirmation = input("Press Enter to continue or type 'no' to skip this record: ")
            if confirmation.lower() == 'no':
                print("Going to next record.")
                break
            try:
                # URL encode the query
                encoded_query = quote(query)
                # Construct the Google search URL
                search_url = f"https://www.google.com/search?q={encoded_query}" # Change URL based on your country. I used .com
                # Open the web browser with the search URL
                webbrowser.open_new_tab(search_url) # my default browser is firefox and it will open a tab if the browser is already on. 
            except Exception as e:
                print(f"An error occurred: {e}")
    if i >= start_record:
        continue # skip the things already finished
